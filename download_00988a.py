import datetime as dt
import pathlib
import re
from typing import Optional, Tuple

import pandas as pd
import requests
from openpyxl import load_workbook

BASE_URL = "https://www.ezmoney.com.tw"
FUND_CODE = "61YTW"  # 00988A 的 fundCode
ETF_CODE = "00988A"

INFO_URL = f"{BASE_URL}/ETF/Fund/Info?FundCode={FUND_CODE}"
EXPORT_URL = f"{BASE_URL}/ETF/Fund/AssetExcelNPOI?fundCode={FUND_CODE}"


def ensure_dir(p: pathlib.Path) -> None:
    p.mkdir(parents=True, exist_ok=True)


def roc_to_ad_yyyymmdd(roc_date_str: str) -> Optional[str]:
    s = str(roc_date_str).strip()
    if not s:
        return None
    s = s.replace("年", "/").replace("月", "/").replace("日", "")
    s = s.replace("-", "/")
    m = re.search(r"(\d{2,3})\s*/\s*(\d{1,2})\s*/\s*(\d{1,2})", s)
    if not m:
        return None
    roc_year = int(m.group(1))
    month = int(m.group(2))
    day = int(m.group(3))
    try:
        d = dt.date(roc_year + 1911, month, day)
    except ValueError:
        return None
    return d.strftime("%Y%m%d")


def extract_data_date_from_xlsx(xlsx_path: pathlib.Path) -> Optional[str]:
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    for r in range(1, 21):
        for c in range(1, 11):
            v = ws.cell(row=r, column=c).value
            if v is None:
                continue
            txt = str(v).strip()
            if "資料日期" in txt:
                after = txt.split("資料日期", 1)[-1].replace("：", ":")
                if ":" in after:
                    after = after.split(":", 1)[-1].strip()
                d = roc_to_ad_yyyymmdd(after)
                if d:
                    wb.close()
                    return d
                v2 = ws.cell(row=r, column=c + 1).value
                d2 = roc_to_ad_yyyymmdd(v2) if v2 is not None else None
                if d2:
                    wb.close()
                    return d2
    wb.close()
    return None


def normalize_colname(s: str) -> str:
    return re.sub(r"\s+", "", str(s)).strip()


def find_header_row(df_raw: pd.DataFrame) -> Optional[int]:
    max_scan = min(60, len(df_raw))
    for i in range(max_scan):
        row = df_raw.iloc[i].astype(str).fillna("").tolist()
        row_join = " ".join([x.strip() for x in row if x and x != "nan"]).strip()
        if not row_join:
            continue
        has_code = any(k in row_join for k in ["代號", "股票代號", "標的代號", "證券代號"])
        has_name = any(k in row_join for k in ["名稱", "股票名稱", "標的名稱", "股名"])
        has_shares = any(k in row_join for k in ["股數", "持股股數", "數量", "持有股數"])
        has_weight = any(k in row_join for k in ["權重", "持股權重", "比重"])
        if has_code and has_name and has_shares:
            return i
        if has_code and has_shares and has_weight:
            return i
    return None


def pick_column(cols, candidates):
    for cand in candidates:
        for col in cols:
            if cand in normalize_colname(col):
                return col
    return None


def to_int_safe(x) -> int:
    if x is None:
        return 0
    s = str(x).strip()
    if s == "" or s.lower() == "nan":
        return 0
    s = s.replace(",", "").replace(" ", "")
    try:
        return int(float(s))
    except ValueError:
        return 0


def to_weight_float(x) -> float:
    if x is None:
        return 0.0
    s = str(x).strip()
    if s == "" or s.lower() == "nan":
        return 0.0
    s = s.replace("%", "").replace(",", "").replace(" ", "")
    try:
        return float(s)
    except ValueError:
        return 0.0


def download_xlsx(session: requests.Session, out_path: pathlib.Path) -> None:
    headers = {
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/130.0 Safari/537.36"
        )
    }
    resp_info = session.get(INFO_URL, headers=headers, timeout=30)
    resp_info.raise_for_status()
    print("[INFO] 打開基金資訊頁成功")

    resp_xlsx = session.get(EXPORT_URL, headers=headers, timeout=60)
    resp_xlsx.raise_for_status()
    print(f"[INFO] 下載 API 回應 Content-Type: {resp_xlsx.headers.get('Content-Type')}")

    out_path.write_bytes(resp_xlsx.content)
    print(f"[OK] Saved XLSX to {out_path}")


def parse_holdings_from_xlsx(xlsx_path: pathlib.Path) -> Tuple[pd.DataFrame, Optional[str]]:
    data_date = extract_data_date_from_xlsx(xlsx_path)
    if data_date:
        print(f"[INFO] data_date = {data_date}")
    else:
        print("[WARN] 無法從檔案內抓到資料日期，將以今天日期做檔名")

    df0 = pd.read_excel(xlsx_path, sheet_name=0, header=None, engine="openpyxl")
    header_row = find_header_row(df0)
    if header_row is None:
        preview = df0.head(25).to_string(index=False)
        raise RuntimeError(
            "找不到表頭列（股票代號/股數/權重）。\n"
            "可能 Excel 格式改了。請檢查 raw 檔案。\n\n"
            f"前 25 列預覽：\n{preview}"
        )

    df = pd.read_excel(xlsx_path, sheet_name=0, header=header_row, engine="openpyxl")
    df.columns = [normalize_colname(c) for c in df.columns]

    code_col = pick_column(df.columns, ["股票代號", "代號", "標的代號", "證券代號"])
    name_col = pick_column(df.columns, ["股票名稱", "名稱", "標的名稱", "股名"])
    shares_col = pick_column(df.columns, ["股數", "持股股數", "數量", "持有股數"])
    weight_col = pick_column(df.columns, ["持股權重", "權重", "比重"])

    if not code_col or not shares_col:
        raise RuntimeError(
            f"找不到必要欄位（股票代號/股數）。\n"
            f"目前欄位：{list(df.columns)}"
        )

    rename_map = {code_col: "code", shares_col: "shares"}
    if name_col:
        rename_map[name_col] = "name"
    if weight_col:
        rename_map[weight_col] = "weight"

    df = df[list(rename_map.keys())].copy().rename(columns=rename_map)
    if "name" not in df.columns:
        df["name"] = ""
    if "weight" not in df.columns:
        df["weight"] = 0.0

    df["code"] = df["code"].astype("string").str.strip()
    df["name"] = df["name"].astype("string").str.strip()
    df["shares"] = df["shares"].apply(to_int_safe)
    df["weight"] = df["weight"].apply(to_weight_float)

    df = df[df["code"].notna()]
    df = df[df["code"].str.len() > 0]
    df = df[~df["code"].str.contains("合計|總計|小計", regex=True, na=False)]

    # 00988A 可能包含海外股票代號，例如 MU US、IFX GY、285A JP，因此允許空格。
    df = df[df["code"].str.match(r"^[0-9A-Za-z.\- ]+$", na=False)]

    df = df.groupby(["code", "name"], as_index=False).agg({"shares": "sum", "weight": "sum"})
    df = df.sort_values("weight", ascending=False).reset_index(drop=True)
    return df[["code", "name", "shares", "weight"]], data_date


def compute_diff(prev_df: pd.DataFrame, curr_df: pd.DataFrame) -> pd.DataFrame:
    prev = prev_df.copy()
    curr = curr_df.copy()
    prev["code"] = prev["code"].astype("string").str.strip()
    curr["code"] = curr["code"].astype("string").str.strip()

    prev = prev.rename(columns={"shares": "prev_shares", "weight": "prev_weight"})
    curr = curr.rename(columns={"shares": "curr_shares", "weight": "curr_weight"})

    merged = prev.merge(curr, on=["code"], how="outer", suffixes=("_prev", "_curr"))
    merged["name"] = merged.get("name_curr", "").fillna("")
    if "name_prev" in merged.columns:
        merged.loc[merged["name"].eq("") | merged["name"].isna(), "name"] = merged["name_prev"].fillna("")

    for col in ["prev_shares", "curr_shares"]:
        merged[col] = merged[col].fillna(0).astype(int)
    for col in ["prev_weight", "curr_weight"]:
        merged[col] = merged[col].fillna(0.0).astype(float)

    merged["shares_delta"] = merged["curr_shares"] - merged["prev_shares"]
    merged["weight_delta"] = merged["curr_weight"] - merged["prev_weight"]

    def status_row(r):
        if r["prev_shares"] == 0 and r["curr_shares"] > 0:
            return "NEW"
        if r["prev_shares"] > 0 and r["curr_shares"] == 0:
            return "OUT"
        if r["shares_delta"] > 0:
            return "UP"
        if r["shares_delta"] < 0:
            return "DOWN"
        if abs(r["weight_delta"]) >= 0.01:
            return "WEIGHT_CHANGE"
        return "SAME"

    merged["status"] = merged.apply(status_row, axis=1)
    order_map = {"NEW": 0, "UP": 1, "DOWN": 2, "OUT": 3, "WEIGHT_CHANGE": 4, "SAME": 5}
    merged["order"] = merged["status"].map(order_map).fillna(99)
    merged = merged.sort_values(["order", "weight_delta", "shares_delta"], ascending=[True, False, False]).drop(columns=["order"])

    return merged[[
        "code",
        "name",
        "prev_shares",
        "curr_shares",
        "shares_delta",
        "prev_weight",
        "curr_weight",
        "weight_delta",
        "status",
    ]].reset_index(drop=True)


def write_summary_markdown(diff_df: pd.DataFrame, out_md: pathlib.Path, data_date: str) -> None:
    def top_rows(status, n=20):
        sub = diff_df[diff_df["status"] == status].copy()
        if status in ("DOWN", "OUT"):
            sub = sub.sort_values(["weight_delta", "shares_delta"], ascending=[True, True])
        else:
            sub = sub.sort_values(["weight_delta", "shares_delta"], ascending=[False, False])
        return sub.head(n)

    lines = []
    lines.append(f"# {ETF_CODE} Holdings Diff ({data_date})\n\n")
    counts = diff_df["status"].value_counts().to_dict()
    lines.append("## Summary\n\n")
    lines.append(
        f"- NEW: {counts.get('NEW',0)} | UP: {counts.get('UP',0)} | DOWN: {counts.get('DOWN',0)} | "
        f"OUT: {counts.get('OUT',0)} | WEIGHT_CHANGE: {counts.get('WEIGHT_CHANGE',0)} | SAME: {counts.get('SAME',0)}\n\n"
    )

    sections = [
        ("NEW", "新增持股"),
        ("UP", "加碼"),
        ("DOWN", "減碼"),
        ("OUT", "出清"),
        ("WEIGHT_CHANGE", "權重變化"),
    ]
    for sec, label in sections:
        sub = top_rows(sec)
        lines.append(f"## {label} ({sec})\n\n")
        if sub.empty:
            lines.append("_None_\n\n")
            continue
        lines.append("| code | name | prev shares | curr shares | shares delta | prev weight | curr weight | weight delta | status |\n")
        lines.append("|---|---|---:|---:|---:|---:|---:|---:|---|\n")
        for _, r in sub.iterrows():
            name = str(r["name"]).replace("|", " ")
            lines.append(
                f"| {r['code']} | {name} | {r['prev_shares']} | {r['curr_shares']} | {r['shares_delta']} | "
                f"{r['prev_weight']:.2f}% | {r['curr_weight']:.2f}% | {r['weight_delta']:.2f}% | {r['status']} |\n"
            )
        lines.append("\n")

    out_md.write_text("".join(lines), encoding="utf-8")


def main():
    base = pathlib.Path("data")
    raw_dir = base / "raw"
    out_dir = base / "out"
    holdings_dir = out_dir / "holdings"
    diff_csv_dir = out_dir / "diff" / "csv"
    diff_md_dir = out_dir / "diff" / "md"

    for d in [raw_dir, holdings_dir, diff_csv_dir, diff_md_dir]:
        ensure_dir(d)

    session = requests.Session()
    tmp_path = raw_dir / f"{ETF_CODE}_portfolio_tmp.xlsx"
    download_xlsx(session, tmp_path)

    holdings_df, data_date = parse_holdings_from_xlsx(tmp_path)
    if not data_date:
        data_date = dt.date.today().strftime("%Y%m%d")

    raw_path = raw_dir / f"{ETF_CODE}_portfolio_{data_date}.xlsx"
    if raw_path.exists():
        tmp_path.unlink(missing_ok=True)
        print(f"[INFO] Raw XLSX already exists: {raw_path}")
    else:
        tmp_path.replace(raw_path)
        print(f"[OK] Raw XLSX moved to: {raw_path}")

    holdings_path = holdings_dir / f"{ETF_CODE}_holdings_{data_date}.csv"
    holdings_df.to_csv(holdings_path, index=False, encoding="utf-8-sig")
    print(f"[OK] Saved standardized holdings to {holdings_path}")

    latest_path = holdings_dir / f"{ETF_CODE}_latest.csv"
    root_latest_path = out_dir / f"{ETF_CODE}_latest.csv"

    if latest_path.exists():
        prev_df = pd.read_csv(latest_path, dtype={"code": "string"})
        if not {"code", "shares", "weight"}.issubset(set(prev_df.columns)):
            print("[WARN] latest.csv 格式不對，將略過 diff。")
        else:
            diff_df = compute_diff(prev_df, holdings_df)
            diff_csv_path = diff_csv_dir / f"{ETF_CODE}_diff_{data_date}.csv"
            diff_md_path = diff_md_dir / f"{ETF_CODE}_diff_{data_date}.md"
            diff_df.to_csv(diff_csv_path, index=False, encoding="utf-8-sig")
            write_summary_markdown(diff_df, diff_md_path, data_date)
            print(f"[OK] Saved diff CSV to {diff_csv_path}")
            print(f"[OK] Saved diff MD to {diff_md_path}")
    else:
        print("[INFO] No previous latest.csv found; diff skipped (first run).")

    holdings_df.to_csv(latest_path, index=False, encoding="utf-8-sig")
    holdings_df.to_csv(root_latest_path, index=False, encoding="utf-8-sig")
    print(f"[OK] Updated latest to {latest_path}")
    print(f"[OK] Updated root latest to {root_latest_path}")


if __name__ == "__main__":
    main()
